import os
import tempfile
from shutil import which

from openpyxl import load_workbook  # type: ignore
import pdfplumber  # type: ignore
import pytest
from fastapi.testclient import TestClient

from processing_tools.main import app

client = TestClient(app)


def test_wkhtmltopdf_convert():
    if not which("wkhtmltopdf"):
        pytest.fail(
            "wkhtmltopdf not found: The `wkhtmltopdf` binary needs to exist and be in the path."
        )

    response = client.post(
        "/html_to_pdf/", json={"url": "https://example.com", "engine": "wkhtmltopdf"}
    )
    assert response.status_code == 200

    assert response.content.startswith(b"%PDF-1.4\n")
    assert len(response.content) > 2048


@pytest.mark.no_deps
def test_wkhtmltopdf_convert_invalid_requests():
    response = client.get("/html_to_pdf/")
    assert response.status_code == 405

    response = client.post("/html_to_pdf/")
    assert response.status_code == 422

    response = client.post(
        "/html_to_pdf/", json={"url": "file:///", "engine": "wkhtmltopdf"}
    )
    assert response.status_code == 422


def test_browserless_convert():
    response = client.post(
        "/html_to_pdf/", json={"url": "https://example.com", "engine": "browserless"}
    )
    assert response.status_code == 200

    assert response.content.startswith(b"%PDF-1.4\n")
    assert len(response.content) > 2048


def test_browserless_convert_testfile():
    response = client.post(
        "/html_to_pdf/",
        json={"url": "http://test_server:8081/test.html", "engine": "browserless"},
    )
    assert response.status_code == 200

    with tempfile.TemporaryDirectory() as tmpdir:

        pdf_path = os.path.join(tmpdir, "test.pdf")

        with open(pdf_path, "wb") as out:
            for chunk in response.iter_content(chunk_size=512):
                out.write(chunk)

        with pdfplumber.open(pdf_path) as pdf:
            first_page = pdf.pages[0]
            assert len(pdf.pages) == 1
            text = first_page.extract_text()
            print(text)
            assert text == "My First Heading\nMy first paragraph."

    assert response.content.startswith(b"%PDF-1.4\n")
    assert len(response.content) > 2048


def test_wkhtmltopdf_convert_localfile():
    response = client.post(
        "/html_to_pdf/",
        json={"url": "http://test_server:8081/test.html", "engine": "wkhtmltopdf"},
    )
    assert response.status_code == 200

    with tempfile.TemporaryDirectory() as tmpdir:

        pdf_path = os.path.join(tmpdir, "test.pdf")

        with open(pdf_path, "wb") as out:
            for chunk in response.iter_content(chunk_size=512):
                out.write(chunk)

        with pdfplumber.open(pdf_path) as pdf:
            first_page = pdf.pages[0]
            assert len(pdf.pages) == 1
            text = first_page.extract_text()
            print(text)
            assert text == "My\tFirst\tHeading\nMy\tfirst\tparagraph."

    assert response.content.startswith(b"%PDF-1.4\n")
    assert len(response.content) > 2048


@pytest.mark.no_deps
def test_browserless_convert_invalid_requests():
    response = client.get("/html_to_pdf/")
    assert response.status_code == 405

    response = client.post("/html_to_pdf/")
    assert response.status_code == 422

    response = client.post(
        "/html_to_pdf/", json={"url": "file:///", "engine": "browserless"}
    )
    assert response.status_code == 422


def test_docx_convert():
    response = client.post(
        "/od_to_pdf/",
        json={"url": "http://test_server:8081/test-word.docx"},
    )
    assert response.status_code == 200

    with tempfile.TemporaryDirectory() as tmpdir:

        pdf_path = os.path.join(tmpdir, "test.pdf")

        with open(pdf_path, "wb") as out:
            for chunk in response.iter_content(chunk_size=512):
                out.write(chunk)

        with pdfplumber.open(pdf_path) as pdf:
            first_page = pdf.pages[0]
            assert len(pdf.pages) == 1
            text = first_page.extract_text()
            print(text)
            assert text == "This is a sample word document.\nHi."

    assert response.content.startswith(b"%PDF-1.5\n")
    assert len(response.content) > 2048


def test_ppt_simple_convert():
    response = client.post(
        "/od_to_pdf/",
        json={"url": "http://test_server:8081/sample-powerpoint.pptx"},
    )
    assert response.status_code == 200

    with tempfile.TemporaryDirectory() as tmpdir:

        pdf_path = os.path.join(tmpdir, "test.pdf")

        with open(pdf_path, "wb") as out:
            for chunk in response.iter_content(chunk_size=512):
                out.write(chunk)

        with pdfplumber.open(pdf_path) as pdf:
            first_page = pdf.pages[0]
            assert len(pdf.pages) == 1
            text = first_page.extract_text()
            print(text)
            assert text == "Here’s a slide\nThis is the text inside the slide. Ok."

    assert response.content.startswith(b"%PDF-1.5\n")
    assert len(response.content) > 2048


def test_ppt_more_complex_convert():
    response = client.post(
        "/od_to_pdf/",
        json={"url": "http://test_server:8081/power-point-background.pptx"},
    )
    assert response.status_code == 200

    with tempfile.TemporaryDirectory() as tmpdir:

        pdf_path = os.path.join(tmpdir, "test.pdf")

        with open(pdf_path, "wb") as out:
            for chunk in response.iter_content(chunk_size=512):
                out.write(chunk)

        with pdfplumber.open(pdf_path) as pdf:
            first_page = pdf.pages[0]
            assert len(pdf.pages) == 1
            text = first_page.extract_text()
            print(text)
            assert text == "Slide with a \nhouse\nHow pretty!"

    assert response.content.startswith(b"%PDF-1.5\n")
    assert len(response.content) > 2048


def test_xls_to_xlsx_convert():
    response = client.post(
        "/xls_to_xlsx/",
        json={"url": "http://test_server:8081/sample.xls"},
    )
    assert response.status_code == 200

    with tempfile.TemporaryDirectory() as tmpdir:

        xlsx_path = os.path.join(tmpdir, "test.xlsx")

        with open(xlsx_path, "wb") as out:
            for chunk in response.iter_content(chunk_size=512):
                out.write(chunk)
        wb = load_workbook(filename=xlsx_path, read_only=True)
        ws = wb.active
        assert ws["A1"].value == "Sample File"
        assert ws["B1"].value == "This is B1"
        wb.close()

    assert response.content.startswith(b"PK\x03\x04\x14\x00")
    assert len(response.content) > 2048
